from telebot import types
import telebot.types
import sentAnalysis as sentiment
import time
from openpyxl import load_workbook

TOKEN = '5878190929:AAGdL1hYqGcuTCjYMOefkx3VQvXV39VHRxs'
bot = telebot.TeleBot(TOKEN)
FEEDBACK = ''
LABEL = ''


@bot.message_handler(commands=["start"])
def welcome(message):
    sticker = open('stickers/hello.webp', 'rb')
    bot.send_sticker(message.chat.id, sticker)

    markup = types.ReplyKeyboardMarkup(row_width=2)
    item1 = types.KeyboardButton("Оценить тональность отзыва")
    item2 = types.KeyboardButton("Слова, влияющие на тональность")
    markup.add(item1, item2)

    bot.send_message(message.chat.id,
                     "Добро пожаловать, {0.first_name}!\nЯ бот для анализа тональности отзывов о высших учебных заведениях РФ.\nОтправь мне свой отзыв и я определю его тональность.".format(
                         message.from_user), parse_mode='html', reply_markup=markup)


@bot.message_handler(func=lambda msg: msg.text == 'Слова, влияющие на тональность')
def get_tonality_words(message: types.Message):
    photo = open('positive_words.png', 'rb')
    bot.send_photo(message.chat.id, photo)


@bot.message_handler(func=lambda msg: msg.text == 'Оценить тональность отзыва')
def get_feedback(message: types.Message):
    bot.send_message(message.chat.id, 'Введите текст вашего отзыва')


@bot.message_handler(content_types=["text"])
def analysisFeedback(message):
    label = sentiment.predict_tonality(message.text)
    add_to_excel(message.text, label, False)
    bot.send_message(message.chat.id, 'Ваш отзыв ' + label)

    markup = types.InlineKeyboardMarkup(row_width=2)

    item1 = types.InlineKeyboardButton("Да", callback_data='yes')
    item2 = types.InlineKeyboardButton("Нет", callback_data='no')

    markup.add(item1, item2)

    bot.send_message(message.chat.id, 'Согласны ли вы с результатом?', reply_markup=markup)


@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    try:
        if call.message:
            if call.data == 'yes':
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="Спасибо, мы учтем ваш ответ.")
            elif call.data == 'no':
                markup = types.InlineKeyboardMarkup(row_width=3)

                item1 = types.InlineKeyboardButton("Негативный", callback_data='negative')
                item2 = types.InlineKeyboardButton("Нейтральный", callback_data='neutral')
                item3 = types.InlineKeyboardButton("Позитивный", callback_data='positive')

                markup.add(item1, item2, item3)

                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="К какому классу вы отнесли бы свой отзыв?", reply_markup=markup)
            elif call.data == 'negative':
                add_to_excel('', "негативный", True)
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="Спасибо, мы учтем ваши исправления.")
            elif call.data == 'neutral':
                add_to_excel('', "нейтральный", True)
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="Спасибо, мы учтем ваши исправления.")
            elif call.data == 'positive':
                add_to_excel('', "позитивный", True)
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="Спасибо, мы учтем ваши исправления.")

    except Exception as e:
        print(repr(e))


def add_to_excel(text, label, isFix):
    wb = load_workbook('feedbacksFromTelegram.xlsx')
    ws = wb['Sheet1']
    if isFix:
        ws.cell(column=2, row=ws.max_row, value=label)
    else:
        ws.cell(column=1, row=ws.max_row + 1, value=text)
        ws.cell(column=2, row=ws.max_row, value=label)

    wb.save(filename='feedbacksFromTelegram.xlsx')


# RUN
bot.polling(none_stop=True)
